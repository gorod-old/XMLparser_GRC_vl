import json
import os

from MessagePack import print_info_msg
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, fills
from openpyxl.utils import get_column_letter


def save_json(json_data, root_folder: str = '', file_name: str = 'result', encoding='utf-8', folder: str = None):
    path = __get_file_path('json', root_folder=root_folder, file_name=file_name, folder=folder)
    print_info_msg(msg=f'save path: {path}')
    if os.path.exists(path):
        os.remove(path)
    with open(path, 'a', encoding=encoding) as file:
        json.dump(json_data, file, indent=4, ensure_ascii=False)


def get_json_data_from_file(path, encoding='utf-8', stream: int = None):
    print_info_msg(msg=f'get data path: {path}', stream=stream)
    json_content = open(path, 'r', encoding=encoding).read()
    json_data = json.loads(json_content)
    return json_data


def __get_file_path(extension: str, root_folder: str = 'result data', file_name: str = 'result', folder: str = None):
    if root_folder is None or root_folder == '':
        print_info_msg(msg=f'no root folder specified for output data, set to: None')
        root_folder = None
    root = os.getcwd() + f'/{root_folder}' if root_folder is not None else os.getcwd()
    if not os.path.exists(root) or not os.path.isdir(root):
        os.mkdir(root)
    folder = '/' + folder if folder is not None else ''
    if folder != '' and (not os.path.exists(root + folder) or not os.path.isdir(root + folder)):
        os.mkdir(root + folder)
    path = root + folder + f'/{file_name}.{extension}'
    return os.path.normpath(path)


def save_xlsx(p_data, file_name: str = 'result', folder: str = None,
              header_map=None, start_time: datetime = None,
              header_height: int = 30, row_height: int = 20, column_width: int = 15,
              col_width_map=None, auto_start: bool = False):
    """Saving data to xlsx file."""
    # hyper = ['№ аукциона', 'Ссылка на выписку из реестра МинПромТорга', 'Ссылка на РУ', 'Файлы']
    if len(p_data) == 0:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист 1"

    if header_map:
        p_data.insert(0, header_map)
    if start_time:
        p_data.append([])
        p_data.append(['время записи: ' + str(start_time)])

    for row, row_data in enumerate(p_data):
        for col, cell_data in enumerate(row_data):
            _ = ws.cell(column=col + 1, row=row + 1, value="{0}".format(cell_data))

    header_color = PatternFill(start_color='f5eeda',
                               end_color='f5eeda',
                               fill_type='solid')
    if col_width_map:
        for col, width in enumerate(col_width_map):
            ws.column_dimensions[get_column_letter(col + 1)].width = width
            ws.cell(1, col + 1).fill = header_color

    # save file
    path = __get_file_path('xlsx', file_name=file_name, folder=folder)
    print_info_msg(msg=f'save path: {path}')
    if os.path.exists(path):
        os.remove(path)
    wb.save(path)
    if auto_start:
        os.startfile(path)


def save_xlsx_sheets(data: dict, file_name: str = 'result', folder: str = None, start_time: datetime = None,
                     header_height: int = 30, row_height: int = 20, column_width: int = 15, auto_start: bool = False):
    """Saving data to xlsx file."""
    wb = Workbook()
    sheets = wb.sheetnames
    ws = wb.get_sheet_by_name(sheets[0])
    i = 0
    for key, tab_data in data.items():
        if i > len(sheets) - 1:
            ws = wb.create_sheet(key)
        elif sheets[i] != key:
            ws = wb.get_sheet_by_name(sheets[i])
            ws.title = key

        for row, row_data in enumerate(tab_data):
            for col, cell_data in enumerate(row_data):
                _ = ws.cell(column=col + 1, row=row + 1, value="{0}".format(cell_data))

        header_color = PatternFill(start_color='f5eeda',
                                   end_color='f5eeda',
                                   fill_type='solid')

        i += 1

    # save file
    path = __get_file_path('xlsx', file_name=file_name, folder=folder)
    print_info_msg(msg=f'save path: {path}')
    if os.path.exists(path):
        os.remove(path)
    wb.save(path)
    if auto_start:
        os.startfile(path)
